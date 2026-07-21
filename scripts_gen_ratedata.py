#!/usr/bin/env python3
"""info_daily.xlsx(DailyRate) -> rate-data.js 생성.
날짜축을 공유하고 시리즈별 값 배열만 담아 파일 크기를 줄인다."""
import openpyxl, json, datetime, re, sys

SRC = sys.argv[1] if len(sys.argv) > 1 else "/root/.claude/uploads/7d083a69-7ba1-5b52-bf27-4b2f5fdd3b3e/54f3f8f0-info_daily.xlsx"
OUT = "/home/user/economic-dashboard/rate-data.js"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["DailyRate"]

NOTA_ROW, DESC_ROW, DATA_START = 7, 9, 10

def fmt_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]

# 날짜(오름차순) 수집
raw = []
for r in range(DATA_START, ws.max_row + 1):
    d = ws.cell(row=r, column=1).value
    if d is None:
        continue
    raw.append((r, fmt_date(d)))
raw_sorted = sorted(raw, key=lambda x: x[1])
dates = [d for _, d in raw_sorted]
rows_in_order = [r for r, _ in raw_sorted]
date_index = {d: i for i, d in enumerate(dates)}

# 라벨/그룹 매핑 -------------------------------------------------------------
def label_and_group(nota):
    n = nota
    # 국채금리
    m = re.match(r"^KTB(.+)$", n)
    if m: return ("국채금리 · 한국", f"한국 국고채 {m.group(1)}", "한국", "%")
    m = re.match(r"^UST(.+)$", n)
    if m:
        mat = m.group(1)
        if mat == "0y": mat = "10y"  # 원본 표기 오류(UST0y=10년) 보정
        return ("국채금리 · 미국", f"미국 국채 {mat}", "미국", "%")
    m = re.match(r"^JPY(\d.*y)$", n)
    if m: return ("국채금리 · 일본", f"일본 국채 {m.group(1)}", "일본", "%")
    m = re.match(r"^GER(.+)$", n)
    if m: return ("국채금리 · 독일", f"독일 국채 {m.group(1)}", "독일", "%")
    m = re.match(r"^AUD(\d.*y)$", n)
    if m: return ("국채금리 · 호주", f"호주 국채 {m.group(1)}", "호주", "%")
    # 통안채
    m = re.match(r"^MSB(.+)$", n)
    if m: return ("통안채", f"통안채 {m.group(1)}", "한국", "%")
    # 기준금리
    pol = {"BOK_PR": ("한국은행 기준금리", "한국"), "FFR": ("미 연준 기준금리", "미국"),
           "BOJ_PR": ("일본은행 정책금리", "일본"), "AUD_PR": ("호주 RBA 현금금리", "호주"),
           "ECB_MRO": ("ECB 주요재융자금리", "유럽"), "ECB_MLF": ("ECB 한계대출금리", "유럽")}
    if n in pol: return ("기준금리", pol[n][0], pol[n][1], "%")
    # 환율
    fx = {"DXY": ("달러인덱스(DXY)", "미국", "pt"), "USDKRW": ("원/달러 환율", "한국", "원"),
          "USDJPY": ("엔/달러 환율", "일본", "엔"), "USDCNY": ("위안/달러 환율", "중국", "위안"),
          "USDEUR": ("유로/달러 환율", "유럽", "")}
    if n in fx: return ("환율", fx[n][0], fx[n][1], fx[n][2])
    # 스왑포인트 (HP)
    m = re.match(r"^(\w+)KRW_HP_(\w+)$", n)
    if m: return ("스왑포인트", f"{m.group(1)}/원 스왑포인트 {m.group(2)}", "한국", "")
    # 원/달러 스왑레이트 (SMB)
    m = re.match(r"^SMB_USDKRW_(\w+)$", n)
    if m: return ("환율/스왑", f"원/달러 스왑레이트 {m.group(1)}", "한국", "%")
    # 미국 크레딧 스프레드
    sp = {"US_IG_SP": "미국 투자등급 회사채 스프레드", "US_HY_SP": "미국 하이일드 회사채 스프레드"}
    if n in sp: return ("해외 크레딧", sp[n], "미국", "bp")
    # CP
    m = re.match(r"^CP_(\w+?)_(\w+)$", n)
    if m: return ("CP(기업어음)", f"CP {m.group(1)} {m.group(2)}", "한국", "%")
    # KTB Index
    m = re.match(r"^KTBIndex_(.+)$", n)
    if m: return ("국고채 지수", f"국고채지수 {m.group(1)}", "한국", "pt")
    # 물가연동 BEI
    m = re.match(r"^(\w+)_BEI(\d+y)$", n)
    if m: return ("기대인플레이션(BEI)", f"{m.group(1)} BEI {m.group(2)}", m.group(1), "%")
    # 미국 크레딧
    cr = {"US_IG_YTM": "미국 투자등급 회사채 YTM", "US_HY_YTM": "미국 하이일드 회사채 YTM",
          "UST_ALL_YTM": "미국 국채 전체 YTM"}
    if n in cr: return ("해외 크레딧", cr[n], "미국", "%")
    # 국내 여전채/은행채/공사채/회사채
    m = re.match(r"^(Public|KDB|Bank|Corp|Card|capital)_(.+)$", n)
    if m:
        kind = {"Public": "공사채", "KDB": "산금채", "Bank": "은행채",
                "Corp": "회사채", "Card": "카드채", "capital": "캐피탈채"}[m.group(1)]
        return ("국내 크레딧", f"{kind} {m.group(2).replace('_',' ')}", "한국", "%")
    # 기타
    return ("기타", n, "", "")

series = []
seen_ids = {}
for c in range(2, ws.max_column + 1):
    nota = ws.cell(row=NOTA_ROW, column=c).value
    if nota is None:
        continue
    nota = str(nota).strip()
    group, name, country, unit = label_and_group(nota)
    sid = re.sub(r"[^A-Za-z0-9]+", "_", nota).strip("_").lower()
    if sid in seen_ids:
        seen_ids[sid] += 1
        sid = f"{sid}_{seen_ids[sid]}"
    else:
        seen_ids[sid] = 0
    values = [None] * len(dates)
    n_valid = 0
    for r in range(DATA_START, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if d is None:
            continue
        key = fmt_date(d)
        idx = date_index.get(key)
        if idx is None:
            continue
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            values[idx] = round(float(v), 4)
            n_valid += 1
    if n_valid == 0:
        continue
    series.append({"id": sid, "group": group, "country": country,
                   "name": name, "unit": unit, "notation": nota, "values": values})

# JS 출력 (values는 null 포함, 컴팩트하게)
def js_num(v):
    if v is None: return "null"
    if v == int(v): return str(int(v))
    return repr(v)

with open(OUT, "w", encoding="utf-8") as f:
    f.write("// 자동 생성 파일 — info_daily.xlsx(DailyRate) 원본에서 추출한 일별 금리·환율·크레딧 시계열.\n")
    f.write(f"// 기간 {dates[0]} ~ {dates[-1]} (영업일 {len(dates)}일), 시리즈 {len(series)}개.\n")
    f.write("// 날짜축(rateData.dates)을 공유하고 각 시리즈는 값 배열만 보관합니다(누락일은 null).\n")
    f.write("const rateData = {\n")
    f.write("  dates: [" + ",".join(f'"{d}"' for d in dates) + "],\n")
    f.write("  series: [\n")
    for s in series:
        vals = ",".join(js_num(v) for v in s["values"])
        f.write("    {"
                f'id:"{s["id"]}",group:"{s["group"]}",country:"{s["country"]}",'
                f'name:"{s["name"]}",unit:"{s["unit"]}",notation:"{s["notation"]}",'
                f"values:[{vals}]" + "},\n")
    f.write("  ],\n};\n")

print(f"dates={len(dates)} series={len(series)} -> {OUT}")
import os
print("size:", round(os.path.getsize(OUT)/1024, 1), "KB")
