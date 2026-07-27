#!/usr/bin/env python3
# ============================================================================
# update-rates.py
# 인포맥스류 일별 금리/환율 엑셀(info_daily.xlsx, 시트 'DailyRate')을 읽어
# rate-data.js(일별 금리·환율·크레딧 시계열)를 "최신 영업일만 이어붙이는"
# 방식으로 갱신합니다. 그리고 블룸버그(bloomberg-data.js)와 겹치는 시리즈를
# 교차검증해 보고합니다.
#
# 사용법:
#   1) 엑셀을 data-imports/info_daily.xlsx 로 저장
#   2) python3 scripts/update-rates.py            # 갱신 + 교차검증
#      python3 scripts/update-rates.py --dry       # 미리보기(파일 안 씀)
#
# 엑셀 형식(DailyRate 시트): 'Notation' 행에 각 열의 코드(KTB10y, UST0y,
#   GER10y ...), 그 아래 데이터행은 [날짜 | 값...]. (날짜는 최신→과거 순도 허용)
#
# 동작(안전 우선 — "최신 영업일만 추가"):
#   - rateData.dates 의 마지막 날짜보다 더 최신인 날짜만 추가합니다.
#   - 각 시리즈는 notation(중복 시 등장 순서)으로 엑셀 열과 매칭해 값을 이어붙입니다.
#   - 엑셀에 없는 시리즈는 새 날짜에 null 을 넣어 날짜축 정렬만 맞춥니다(값 보존).
#   - 과거 값·메타데이터(id/group/name/unit/notation)는 절대 바꾸지 않습니다.
#
# 교차검증: 아래 CROSSCHECK 쌍(rate-data notation ↔ bloomberg key)의 최근 공통
#   시점 값을 비교해, 차이가 THRESHOLD 초과면 ⚠️ 로 표시합니다.
#
# 필요: openpyxl, node(설치돼 있음)
# ============================================================================
import sys, os, re, json, subprocess, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RATE_JS = os.path.join(ROOT, "rate-data.js")
BBG_JS = os.path.join(ROOT, "bloomberg-data.js")
INPUT = os.path.join(ROOT, "data-imports", "info_daily.xlsx")
SHEET = "DailyRate"
DRY = "--dry" in sys.argv

# rate-data notation ↔ bloomberg daily key (겹치는 지표 교차검증)
CROSSCHECK = [
    ("KTB10y", "gvsk10yr_index", "한국 10Y"),
    ("UST0y", "usgg10yr_index", "미국 10Y"),
    ("JPY10y", "gjgb10_index", "일본 10Y"),
    ("GER10y", "gdbr10_index", "독일 10Y"),
    ("AUD10y", "gacgb10_index", "호주 10Y"),
    ("USDKRW", "usdkrw_index", "원/달러"),
    ("DXY", "dxy_index", "달러인덱스"),
]
THRESHOLD = {"원/달러": 5.0, "달러인덱스": 0.5}  # 그 외 금리는 0.15%p
DEFAULT_THRESHOLD = 0.15


def node_json(path, var):
    js = (
        'const fs=require("fs");let s=fs.readFileSync(%r,"utf8").replace(/^\\/\\/.*$/gm,"");'
        's=s.replace(/^\\s*const %s\\s*=\\s*/m,"").replace(/;\\s*$/,"");'
        'process.stdout.write(JSON.stringify(eval("("+s+")")));' % (path, var)
    )
    r = subprocess.run(["node", "-e", js], capture_output=True, text=True)
    if r.returncode != 0:
        sys.exit(f"{path} 읽기 실패:\n{r.stderr}")
    return json.loads(r.stdout)


def is_bad(v):
    return v is None or (isinstance(v, str) and ("#N/A" in v or v.strip() == ""))


def num(v):
    if is_bad(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def norm_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    return None


def parse_xlsx():
    import openpyxl
    wb = openpyxl.load_workbook(INPUT, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"'{SHEET}' 시트를 찾지 못했습니다. (시트: {wb.sheetnames})")
    rows = list(wb[SHEET].iter_rows(values_only=True))
    notrow = next((i for i, r in enumerate(rows[:15]) if r and r[0] == "Notation"), None)
    if notrow is None:
        sys.exit("'Notation' 행을 찾지 못했습니다.")
    nrow = rows[notrow]
    cols = [(ci, str(nrow[ci]).strip()) for ci in range(1, len(nrow)) if not is_bad(nrow[ci])]
    # 데이터: col0 이 날짜인 행
    colseries = {ci: {} for ci, _ in cols}
    for r in rows:
        if not r:
            continue
        d = norm_date(r[0])
        if not d:
            continue
        for ci, _ in cols:
            v = num(r[ci]) if ci < len(r) else None
            if v is not None:
                colseries[ci][d] = v
    return cols, colseries


def main():
    if not os.path.exists(INPUT):
        sys.exit(f"입력 파일이 없습니다: {INPUT}\n엑셀을 이 경로에 info_daily.xlsx 로 저장하세요.")
    try:
        import openpyxl  # noqa
    except ImportError:
        sys.exit("openpyxl 가 필요합니다:  pip install openpyxl")

    cols, colseries = parse_xlsx()
    rd = node_json(RATE_JS, "rateData")
    dates = rd["dates"]
    series = rd["series"]
    cur_last = dates[-1]

    # 새 날짜(현재 마지막보다 최신) 오름차순
    all_xdates = set()
    for _, s in colseries.items():
        all_xdates.update(s.keys())
    new_dates = sorted(d for d in all_xdates if d > cur_last)

    # notation(등장순) ↔ 엑셀 열 매칭용: 엑셀 열을 notation별 큐로
    from collections import defaultdict, deque
    xqueue = defaultdict(deque)
    for ci, notation in cols:
        xqueue[notation].append(ci)
    # 시리즈별 매칭 열 결정(등장 순서대로 소비)
    counter = defaultdict(int)
    match_col = {}   # series_index -> xlsx col
    for si, s in enumerate(series):
        notation = s.get("notation")
        q = xqueue.get(notation)
        if q:
            k = counter[notation]
            if k < len(q):
                match_col[si] = q[k]
                counter[notation] += 1
    matched = len(match_col)

    # 새 날짜만큼 값 이어붙이기
    if new_dates:
        for d in new_dates:
            dates.append(d)
        for si, s in enumerate(series):
            ci = match_col.get(si)
            src = colseries.get(ci, {}) if ci is not None else {}
            for d in new_dates:
                s["values"].append(src.get(d))  # 없으면 None(null)

    # ---- 교차검증 (rate-data ↔ bloomberg) ----
    bbg = node_json(BBG_JS, "bloombergData")
    notation_series = {}
    for s in series:
        notation_series.setdefault(s.get("notation"), s)
    xlog = []
    for notation, bkey, label in CROSSCHECK:
        s = notation_series.get(notation)
        bd = bbg.get("daily", {}).get(bkey)
        if not s or not bd:
            xlog.append(f"  {label:8s}: (한쪽 없음 — rate={bool(s)}, bbg={bool(bd)})")
            continue
        # rate-data 최신 (date,val)
        rlast = None
        for d, v in zip(dates, s["values"]):
            if v is not None:
                rlast = (d, v)
        bmap = {d: v for d, v in bd["series"]}
        blast = bd["series"][-1] if bd["series"] else None
        # 공통 최신 시점: rate 최신일 이하의 bbg 값(주간이라 근접일)
        common = None
        if rlast:
            cand = [d for d in bmap if d <= rlast[0]]
            if cand:
                cd = max(cand)
                common = (cd, bmap[cd], rlast[1])
        if common:
            cd, bv, rv = common
            th = THRESHOLD.get(label, DEFAULT_THRESHOLD)
            diff = abs(bv - rv)
            flag = "⚠️" if diff > th else "✅"
            xlog.append(f"  {flag} {label:8s}: rate={rv:g}({rlast[0]}) vs bbg={bv:g}({cd})  Δ={diff:.3g}")
        else:
            xlog.append(f"  {label:8s}: rate 최신={rlast}, bbg 최신={blast}")

    print(f"엑셀 열 수       : {len(cols)}")
    print(f"매칭된 시리즈    : {matched}/{len(series)}")
    print(f"기존 마지막 날짜 : {cur_last}")
    print(f"신규 영업일      : {len(new_dates)}개  {new_dates if new_dates else ''}")
    print("교차검증 (rate-data ↔ bloomberg, 최근 공통시점):")
    print("\n".join(xlog))

    if DRY:
        print("\n[--dry] 파일을 쓰지 않았습니다.")
        return
    if not new_dates:
        print("\n신규 영업일이 없어 rate-data.js 변경 없음 (이미 최신).")
        return

    # ---- rate-data.js 재작성 ----
    n_days = len(dates)
    n_series = len(series)
    header = (
        "// 자동 생성 파일 — info_daily.xlsx(DailyRate) 원본에서 추출한 일별 금리·환율·크레딧 시계열.\n"
        f"// 기간 {dates[0]} ~ {dates[-1]} (영업일 {n_days}일), 시리즈 {n_series}개.\n"
        "// 날짜축(rateData.dates)을 공유하고 각 시리즈는 값 배열만 보관합니다(누락일은 null).\n"
        "// scripts/update-rates.py 로 갱신합니다.\n"
    )
    out = [header, "const rateData = {"]
    out.append('  "dates": ' + json.dumps(dates) + ",")
    out.append('  "series": [')
    for i, s in enumerate(series):
        comma = "," if i < n_series - 1 else ""
        out.append("    " + json.dumps(s, ensure_ascii=False, separators=(",", ":")) + comma)
    out.append("  ]")
    out.append("};")
    open(RATE_JS, "w", encoding="utf-8").write("\n".join(out) + "\n")
    print(f"\n✅ rate-data.js 갱신 완료 ({cur_last} → {dates[-1]}).")


if __name__ == "__main__":
    main()
