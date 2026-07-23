#!/usr/bin/env python3
# ============================================================================
# update-calendar.py
# 경제 캘린더 엑셀(인베스팅닷컴류 export)을 읽어 data.js의
# <<CALENDAR_RAW_START>> ~ <<CALENDAR_RAW_END>> 사이(raw 캘린더 이벤트)를
# 자동으로 다시 씁니다.
#
# 사용법:
#   1) 엑셀을 data-imports/calendar.xlsx 로 저장 (아래 INPUT 경로)
#   2) python3 scripts/update-calendar.py
#   3) 필요하면 --dry 로 미리보기만: python3 scripts/update-calendar.py --dry
#
# 엑셀 컬럼(1행 헤더): 날짜 | 시간 | 통화 | 국가 | 발표월 | 표시 | 중요도 |
#                     실제값 | 예상값 | Forecast | 이전값
#
# 동작:
#   - 엑셀에 담긴 날짜 구간[min~max]만 교체합니다. 그 밖의 기존 이벤트는 보존.
#   - 중요도 3→상, 2→중, 1→하 로 변환.
#   - 국가명 정규화(대한민국→한국, 유로 지역→유럽 등).
#   - 지표사전과 연결된 캘린더 이벤트(indicatorId)와 같은 날·같은 국가·같은
#     개념이면 중복이므로 엑셀 쪽 행을 제외(캘린더에 두 번 뜨지 않게).
#   - cal_ id 를 1부터 다시 매깁니다.
#
# 필요: pip install openpyxl (실행 환경엔 이미 설치돼 있음)
# ============================================================================
import sys, os, re

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_JS = os.path.join(ROOT, "data.js")
INPUT = os.path.join(ROOT, "data-imports", "calendar.xlsx")
START = "// <<CALENDAR_RAW_START>>"
END = "// <<CALENDAR_RAW_END>>"

DRY = "--dry" in sys.argv

# 국가명 정규화 (엑셀 표기 → 대시보드 표기)
COUNTRY = {
    "미국": "미국", "대한민국": "한국", "한국": "한국", "일본": "일본",
    "유로 지역": "유럽", "유로지역": "유럽", "유럽": "유럽", "독일": "독일",
    "영국": "영국", "중국": "중국", "호주": "호주",
}
IMPORTANCE = {"3": "상", "2": "중", "1": "하"}

# 지표 id 접두어 → 국가 (중복판정용)
PREFIX_COUNTRY = {"us": "미국", "kr": "한국", "jp": "일본", "eu": "유럽",
                  "de": "유럽", "au": "호주", "cn": "중국", "uk": "영국"}

# 지표 id → 중복판정 키워드 { include: 하나라도 포함하면 후보, exclude: 있으면 제외 }
# (같은 날짜에 그 지표의 캘린더 이벤트가 있을 때만 적용되므로 과도 제외 위험이 낮음)
CONCEPT = {
    "us_cpi":        (["CPI", "소비자물가"], ["근원", "Core"]),
    "us_core_cpi":   (["근원 CPI", "근원 소비자물가", "Core CPI"], []),
    "us_pce":        (["PCE", "개인소비지출"], []),
    "us_nfp":        (["비농업"], []),
    "us_unemployment": (["실업률"], []),
    "us_ism_mfg":    (["ISM 제조업 PMI"], []),
    "us_ism_svc":    (["ISM 서비스업 PMI", "ISM 비제조업"], []),
    "us_retail_sales": (["소매판매"], []),
    "us_durable_goods": (["내구재"], []),
    "us_housing_starts": (["주택착공", "주택 착공"], []),
    "us_cb_consumer": (["컨퍼런스보드", "CB 소비자"], []),
    "us_umich_consumer": (["미시간"], []),
    "us_fomc_meeting": (["FOMC", "연방기금", "Fed 금리", "금리 결정", "금리결정"], ["의사록", "회의록", "연설"]),
    "us_fomc_minutes": (["FOMC 의사록", "의사록", "회의록"], []),
    "kr_cpi":        (["CPI", "소비자물가"], ["근원"]),
    "kr_ip":         (["산업생산"], []),
    "kr_trade":      (["무역수지"], []),
    "kr_bok_meeting": (["기준금리", "금리 결정", "금리결정"], ["의사록"]),
    "jp_cpi":        (["CPI", "소비자물가"], ["근원"]),
    "jp_trade":      (["무역수지"], []),
    "jp_unemployment": (["실업률"], []),
    "jp_boj_meeting": (["BoJ 금리", "일본은행 금리", "금리 결정", "금리결정"], ["회의록", "의사록"]),
    "jp_boj_minutes": (["회의록", "의사록"], []),
    "eu_cpi":        (["CPI", "소비자물가"], ["근원"]),
    "eu_ecb_meeting": (["ECB", "예금 금리", "기준금리", "금리결정"], ["의사록", "연설"]),
    "eu_ecb_minutes": (["의사록", "회의록"], []),
    "eu_pmi_mfg":    (["제조업 PMI"], []),
    "de_ifo":        (["IFO", "Ifo"], []),
    "au_cpi":        (["CPI", "소비자물가"], ["근원"]),
    "au_unemployment": (["실업률"], []),
    "au_retail_sales": (["소매판매"], []),
    "au_rba_meeting": (["RBA 금리", "현금금리", "금리 결정", "금리결정"], ["기자", "회견", "의사록"]),
    "au_rba_minutes": (["의사록", "회의록"], []),
    "cn_trade":      (["무역수지"], []),
    "cn_pmi_official": (["NBS 제조업 PMI", "NBS 비제조업 PMI"], []),
    "cn_retail_sales": (["소매판매"], []),
    "cn_pboc_lpr":   (["LPR", "대출우대금리"], []),
}


def id_country(indid):
    return PREFIX_COUNTRY.get(indid.split("_", 1)[0], "?")


def load_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        sys.exit("openpyxl 가 필요합니다:  pip install openpyxl")


def parse_existing_raw(block_text):
    events = []
    for line in block_text.split("\n"):
        if "raw: true" not in line:
            continue
        d = {}
        for k, v in re.findall(r'(\w+):\s*"((?:[^"\\]|\\.)*)"', line):
            d[k] = v
        if d.get("date"):
            events.append(d)
    return events


def indicator_events(src):
    """calendarEvents 안의 {date, indicatorId} 목록 → set[(date, indicatorId)]"""
    m = re.search(r"const calendarEvents = \[(.*?)\n\];", src, re.S)
    body = m.group(1) if m else src
    out = set()
    for line in body.split("\n"):
        dm = re.search(r'date:\s*"(2\d{3}-\d\d-\d\d)".*indicatorId:\s*"(\w+)"', line)
        if dm:
            out.add((dm.group(1), dm.group(2)))
    return out


def is_dup(date, country, name, ind_by_date):
    """같은 날짜에 있는 지표 이벤트와 개념이 겹치면 True"""
    for indid in ind_by_date.get(date, []):
        if id_country(indid) != country:
            continue
        inc, exc = CONCEPT.get(indid, (None, None))
        if not inc:
            continue
        if any(k in name for k in inc) and not any(k in name for k in exc):
            return True
    return False


def esc(s):
    return s.replace("\\", "\\\\").replace('"', '\\"')


def emit(o):
    parts = [f'id: "{o["id"]}"', f'date: "{o["date"]}"', f'timeKST: "{esc(o.get("timeKST",""))}"',
             f'country: "{esc(o["country"])}"', f'name: "{esc(o["name"])}"', f'importance: "{o["importance"]}"']
    for f in ("actual", "consensus", "forecast", "previous"):
        if o.get(f):
            parts.append(f'{f}: "{esc(o[f])}"')
    parts.append("raw: true")
    return "  { " + ", ".join(parts) + " },"


def main():
    if not os.path.exists(INPUT):
        sys.exit(f"입력 파일이 없습니다: {INPUT}\n엑셀을 이 경로에 calendar.xlsx 로 저장하세요.")
    openpyxl = load_openpyxl()

    src = open(DATA_JS, encoding="utf-8").read()
    if START not in src or END not in src:
        sys.exit("data.js 에 캘린더 마커(<<CALENDAR_RAW_START/END>>)가 없습니다.")
    head, rest = src.split(START, 1)
    block, tail = rest.split(END, 1)

    existing = parse_existing_raw(block)

    # --- 엑셀 파싱 ---
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    ws = wb[wb.sheetnames[0]]
    xrows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]

    def clean(v):
        return "" if v is None else str(v).strip()

    def norm_date(v):
        return str(v).replace("/", "-")[:10]

    xdates = [norm_date(r[0]) for r in xrows]
    wmin, wmax = min(xdates), max(xdates)

    ind_by_date = {}
    for d, i in indicator_events(src):
        ind_by_date.setdefault(d, []).append(i)

    new = []
    dropped_dup = 0
    for r in xrows:
        date = norm_date(r[0])
        time = clean(r[1])
        country = COUNTRY.get(clean(r[3]), clean(r[3]))
        name = clean(r[5])
        imp = IMPORTANCE.get(clean(r[6]), "중")
        actual, cons, fc, prev = clean(r[7]), clean(r[8]), clean(r[9]), clean(r[10])
        if not name:
            continue
        if is_dup(date, country, name, ind_by_date):
            dropped_dup += 1
            continue
        o = {"date": date, "timeKST": time, "country": country, "name": name, "importance": imp, "raw": True}
        if actual: o["actual"] = actual
        if cons: o["consensus"] = cons
        if fc: o["forecast"] = fc
        if prev: o["previous"] = prev
        new.append(o)

    # 엑셀 구간 밖의 기존 이벤트는 그대로 유지
    kept = [e for e in existing if not (wmin <= e.get("date", "") <= wmax)]
    combined = kept + new
    combined.sort(key=lambda e: (e["date"], e.get("timeKST", "")))
    for i, o in enumerate(combined, 1):
        o["id"] = f"cal_{i}"

    print(f"엑셀 구간        : {wmin} ~ {wmax}  (행 {len(xrows)})")
    print(f"중복 제외        : {dropped_dup}건 (지표사전 연결 이벤트와 겹침)")
    print(f"구간 밖 기존 유지: {len(kept)}건")
    print(f"엑셀 신규        : {len(new)}건")
    print(f"최종 raw 이벤트  : {len(combined)}건")

    new_block = "\n" + "\n".join(emit(o) for o in combined) + "\n  "
    out = head + START + new_block + END + tail

    if DRY:
        print("\n[--dry] 파일을 쓰지 않았습니다. 미리보기 8월 상(上) 샘플:")
        for o in combined:
            if o["date"].startswith(wmax[:7]) and o["importance"] == "상":
                print("  " + emit(o)[:120])
        return

    open(DATA_JS, "w", encoding="utf-8").write(out)
    print(f"\n✅ data.js 갱신 완료.")


if __name__ == "__main__":
    main()
