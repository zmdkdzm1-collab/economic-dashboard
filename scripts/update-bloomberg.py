#!/usr/bin/env python3
# ============================================================================
# update-bloomberg.py
# 블룸버그 매크로 데이터 엑셀(고정 템플릿)을 읽어 bloomberg-data.js를
# "최신값만 이어붙이는" 방식으로 갱신합니다. (과거 이력·구조는 그대로 보존)
#
# 사용법:
#   1) 블룸버그 엑셀을 data-imports/bloomberg.xlsx 로 저장
#   2) python3 scripts/update-bloomberg.py            # 갱신
#      python3 scripts/update-bloomberg.py --dry       # 미리보기(파일 안 씀)
#
# 동작 원칙(안전 우선):
#   - 각 시리즈의 티커를 엑셀에서 찾아, 현재 저장된 마지막 날짜보다 "더 최신인
#     데이터만" 이어붙입니다. 기존 과거값/라벨/단위는 절대 바꾸지 않습니다.
#   - 월별 지표: 최신치(col=최신치)를 series에 추가, releases(최근 24개월,
#     {date,actual=최신치,survey=시장예상치,initial=초기발표치})를 롤포워드,
#     nextRelease·survey_latest 갱신.
#   - 일별(금융) 지표: 주간(일요일 기준) 다운샘플로 신규 구간만 이어붙이고 asOf 갱신.
#   - 엑셀에서 티커를 못 찾은 시리즈는 건드리지 않고 보고만 합니다.
#   - 엑셀에만 있고 대시보드에 없는 "새 지표"는 자동 생성하지 않습니다(수동).
#
# 엑셀 형식: 각 데이터 시트의 헤더에 '티커'/'단위'/'지표'/'차기 발표일시' 행이 있고,
#   각 지표는 [최신치|초기 발표치|시장 예상치] (금융은 [종가]) 열로 구성됩니다.
#   (열마다 '티커' 행에 블룸버그 티커가 라벨로 있어 위치가 바뀌어도 안전)
#
# 필요: pip install openpyxl
# ============================================================================
import sys, os, re, json
from datetime import datetime, timedelta

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_JS = os.path.join(ROOT, "bloomberg-data.js")
INPUT = os.path.join(ROOT, "data-imports", "bloomberg.xlsx")
DRY = "--dry" in sys.argv

DATA_SHEETS = ["한국", "미국", "유중일호",
               "한국_Daily 금융지표", "미국_Daily 금융지표", "유중일호_Daily 금융지표",
               "스프레드"]
FIELD_LABELS = {"최신치", "초기 발표치", "시장 예상치", "종가"}


def keyof(ticker):
    return re.sub(r"\s+", "_", str(ticker).strip().lower())


def norm_date(v):
    if hasattr(v, "year"):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    s = str(v)
    m = re.match(r"(\d{4})[/-](\d{2})[/-](\d{2})", s)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else None


def is_bad(v):
    return v is None or (isinstance(v, str) and ("#N/A" in v or v.strip() == ""))


def num(v):
    if is_bad(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_sheet(rows):
    """시트 rows → { ticker_key: {unit,name,nextRelease, fields:{field:col}, series:{field:[(date,val)]}} }

    필드는 '티커' 행에서 같은 티커가 연속으로 반복되는 열 묶음으로 판별한다:
      3열 묶음 = 매크로 [최신치, 초기 발표치, 시장 예상치]
      1열      = 금융 [종가]  (일별 종가/수준값)
    (시트 상단의 다른 설정 헤더 블록에 흔들리지 않도록 field 행에 의존하지 않음)
    """
    def find_label(label):
        for ri, row in enumerate(rows):
            if len(row) > 1 and row[1] == label:
                return ri
        return None

    r_tk = find_label("티커")
    if r_tk is None:
        return {}
    trow = rows[r_tk]
    urow = rows[find_label("단위")] if find_label("단위") is not None else []
    nrow = rows[find_label("지표")] if find_label("지표") is not None else []
    xrow = rows[find_label("차기 발표일시")] if find_label("차기 발표일시") is not None else []

    def cell(row, ci):
        return row[ci] if row and ci < len(row) and not is_bad(row[ci]) else None

    meta = {}
    n = len(trow)
    ci = 2
    while ci < n:
        tk = trow[ci]
        if is_bad(tk):
            ci += 1
            continue
        grp = [ci]
        cj = ci + 1
        while cj < n and trow[cj] == tk:
            grp.append(cj)
            cj += 1
        if len(grp) >= 3:
            fields = {"최신치": grp[0], "초기 발표치": grp[1], "시장 예상치": grp[2]}
        elif len(grp) == 2:
            fields = {"최신치": grp[0], "초기 발표치": grp[1]}
        else:
            fields = {"종가": grp[0]}
        k = keyof(tk)
        nr = cell(xrow, grp[0])
        meta.setdefault(k, {"ticker": str(tk).strip(), "unit": cell(urow, grp[0]),
                            "name": cell(nrow, grp[0]), "nextRelease": (str(nr) if nr else None),
                            "fields": fields, "series": {}})
        ci = cj

    for row in rows[r_tk + 1:]:
        if len(row) < 2:
            continue
        d = norm_date(row[1]) if row[1] is not None and hasattr(row[1], "year") else None
        if not d:
            continue
        for m in meta.values():
            for field, c in m["fields"].items():
                v = num(row[c]) if c < len(row) else None
                if v is not None:
                    m["series"].setdefault(field, []).append((d, v))
    for m in meta.values():
        for f in m["series"]:
            m["series"][f].sort()
    return meta


def daily_value_series(m):
    for f in ("종가", "최신치", "PX_LAST"):
        if f in m["series"] and m["series"][f]:
            return m["series"][f]
    # 라벨이 없으면 유일한 시리즈 사용
    return next(iter(m["series"].values()), [])


def weekly_new_points(xseries, after_date, upto):
    """after_date(제외) ~ upto 까지 일요일 그리드로 신규 주간 포인트 생성"""
    if not xseries:
        return []
    by_date = xseries  # sorted (date, val)
    def val_on_or_before(dstr):
        lo, hi, res = 0, len(by_date) - 1, None
        while lo <= hi:
            mid = (lo + hi) // 2
            if by_date[mid][0] <= dstr:
                res = by_date[mid][1]; lo = mid + 1
            else:
                hi = mid - 1
        return res
    start = datetime.strptime(after_date, "%Y-%m-%d").date()
    # 다음 일요일
    s = start + timedelta(days=(6 - start.weekday()) % 7 or 7)
    end = datetime.strptime(upto, "%Y-%m-%d").date()
    out = []
    while s <= end:
        ds = s.isoformat()
        v = val_on_or_before(ds)
        if v is not None:
            out.append([ds, v])
        s += timedelta(days=7)
    # 마지막 asOf 포인트
    if upto > after_date and (not out or out[-1][0] != upto):
        lv = val_on_or_before(upto)
        if lv is not None:
            out.append([upto, lv])
    return out


def main():
    if not os.path.exists(INPUT):
        sys.exit(f"입력 파일이 없습니다: {INPUT}\n블룸버그 엑셀을 이 경로에 bloomberg.xlsx 로 저장하세요.")
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl 가 필요합니다:  pip install openpyxl")

    # 현재 bloomberg-data.js 로드 (주석 제거 후 JSON 파싱)
    raw = open(DATA_JS, encoding="utf-8").read()
    header = raw.split("\n", 1)[0] if raw.startswith("//") else "// Bloomberg 업로드 데이터."
    body = re.sub(r"^//.*\n", "", raw)
    body = re.sub(r"^const bloombergData\s*=\s*", "", body).strip().rstrip(";")
    cur = json.loads(body)

    wb = openpyxl.load_workbook(INPUT, data_only=True, read_only=True)
    tick = {}
    for name in DATA_SHEETS:
        if name not in wb.sheetnames:
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        for k, m in parse_sheet(rows).items():
            tick.setdefault(k, m)

    changed = 0
    missing = []
    max_asof = cur.get("asOf", "")

    # --- 월별 ---
    for key, obj in cur["monthly"].items():
        tk = keyof(obj["ticker"])
        m = tick.get(tk)
        latest = (m["series"].get("최신치") or m["series"].get("종가")) if m else None
        if not latest:
            missing.append(key)
            continue
        surv = dict(m["series"].get("시장 예상치", []))
        init = dict(m["series"].get("초기 발표치", []))
        cur_last = obj["series"][-1][0]
        new = [(d, v) for d, v in latest if d > cur_last]
        if new:
            for d, v in new:
                obj["series"].append([d, v])
                rel = {"date": d, "actual": v}
                if d in surv: rel["survey"] = surv[d]
                if d in init: rel["initial"] = init[d]
                obj["releases"].append(rel)
            obj["releases"] = obj["releases"][-24:]
            changed += 1
        if m.get("nextRelease"):
            obj["nextRelease"] = m["nextRelease"]
        if surv:
            obj["survey_latest"] = surv[max(surv)]

    # --- 일별 ---
    for key, obj in cur["daily"].items():
        tk = keyof(obj["ticker"])
        m = tick.get(tk)
        if not m:
            missing.append(key)
            continue
        xs = daily_value_series(m)
        if not xs:
            missing.append(key)
            continue
        cur_last = obj["series"][-1][0]
        xmax = xs[-1][0]
        if xmax > max_asof:
            max_asof = xmax
        new = weekly_new_points(xs, cur_last, xmax)
        if new:
            obj["series"].extend(new)
            changed += 1

    cur["asOf"] = max_asof

    print(f"엑셀 티커 수     : {len(tick)}")
    print(f"갱신된 시리즈    : {changed}개 (신규 데이터가 있는 것만)")
    print(f"티커 못 찾음     : {len(missing)}개  {missing if missing else ''}")
    print(f"asOf             : {cur['asOf']}")

    if DRY:
        print("\n[--dry] 파일을 쓰지 않았습니다.")
        return
    if changed == 0:
        print("\n신규 데이터가 없어 변경 없음 (이미 최신).")
        return

    out = header.rstrip("\n") + "\n" + "const bloombergData = " + json.dumps(cur, ensure_ascii=False) + ";\n"
    open(DATA_JS, "w", encoding="utf-8").write(out)
    print("\n✅ bloomberg-data.js 갱신 완료.")


if __name__ == "__main__":
    main()
